import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def create_excel_report(recibos_data):
    """
    Crea un archivo Excel con:
    - Formato: A (Nombre), D (Sueldo), E (Adelanto), F (Pagos), G (Saldo)
    - Estética: Idéntica a la imagen, con cabeceras grandes y borde exterior.
    """
    
    # Mapeo manual para meses en español
    meses_es = [
        "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", 
        "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"
    ]
    
    now = datetime.now()
    month_name = meses_es[now.month - 1]
    year = now.year

    try:
        # 2. Crear Workbook y Hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = month_name

        # 3. Definir Estilos
        font_bold = Font(bold=True)
        font_header = Font(bold=True, size=12) 
        font_bold_red = Font(bold=True, color="FF0000")
        font_red = Font(color="FF0000")
        
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        subheader_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        center_align = Alignment(horizontal='center', vertical='center')
        
        thin_side = Side(style='thin')
        medium_side = Side(style='medium') # Borde más grueso
        
        thin_border = Border(
            left=thin_side, 
            right=thin_side, 
            top=thin_side, 
            bottom=thin_side
        )
        
        currency_format = '"$" #,##0.00'

        # 4. Cabecera Principal (Mes y Año)
        ws['D1'] = f"{month_name} {year}"
        ws.merge_cells('D1:G1')
        header_cell = ws['D1']
        header_cell.font = font_bold
        header_cell.fill = header_fill
        header_cell.alignment = center_align
        # Aplicar borde delgado a la cabecera del mes
        for col_idx in range(4, 8):
            ws.cell(row=1, column=col_idx).border = thin_border

        # 5. Cabeceras de Columnas (Fila 3)
        ws['A3'] = 'NOMBRE Y APELLIDO'
        ws.merge_cells('A3:C3')
        for col_idx in range(1, 4):
            cell = ws.cell(row=3, column=col_idx)
            cell.font = font_header 
            cell.fill = subheader_fill
            cell.alignment = center_align
            
        headers = {'D3': 'SUELDO', 'E3': 'ADELANTO', 'F3': 'PAGOS', 'G3': 'SALDO'}
        for cell_ref, value in headers.items():
            cell = ws[cell_ref]
            cell.value = value
            cell.font = font_header 
            cell.fill = subheader_fill
            cell.alignment = center_align

        # 6. Ajustar Ancho de Columnas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 18

        # 7. Llenar Datos
        current_row = 4
        
        for recibo in recibos_data:
            nombre_completo = f"{recibo.get('apellido', 'N/A')}, {recibo.get('nombre', 'N/A')}".upper()
            sueldo = recibo.get('sueldo', 0)
            
            ws[f'A{current_row}'] = nombre_completo
            
            cell_sueldo = ws[f'D{current_row}']
            cell_sueldo.value = sueldo
            cell_sueldo.number_format = currency_format
            
            cell_adelanto = ws[f'E{current_row}']
            cell_adelanto.font = font_red
            cell_adelanto.number_format = currency_format

            cell_pagos = ws[f'F{current_row}']
            cell_pagos.font = font_red
            cell_pagos.number_format = currency_format
            
            cell_saldo = ws[f'G{current_row}']
            cell_saldo.value = f"=D{current_row}-E{current_row}-F{current_row}"
            cell_saldo.font = font_bold
            cell_saldo.number_format = currency_format
            
            # Aplicar bordes DELGADOS (grid interno)
            for col_idx in range(1, 8):
                 ws.cell(row=current_row, column=col_idx).border = thin_border
            
            current_row += 1

        # 8. Fila de TOTALES
        total_row = current_row
        last_data_row = current_row - 1
        
        ws[f'A{total_row}'] = "TOTALES"
        ws[f'A{total_row}'].font = font_bold
        
        cell_total_sueldo = ws[f'D{total_row}']
        cell_total_sueldo.value = f"=SUM(D4:D{last_data_row})"
        cell_total_sueldo.font = font_bold
        cell_total_sueldo.number_format = currency_format

        for letter in ['E', 'F', 'G']:
            cell = ws[f'{letter}{total_row}']
            cell.value = f"=SUM({letter}4:{letter}{last_data_row})"
            if letter != 'G':
                cell.font = font_bold_red
            else:
                cell.font = font_bold
            cell.number_format = currency_format

        # Aplicar bordes DELGADOS (grid interno) a la fila de TOTALES
        for col_idx in range(1, 8):
             ws.cell(row=total_row, column=col_idx).border = thin_border

        # Define el rango de toda la tabla (desde cabeceras a totales)
        start_row = 3
        end_row = total_row
        start_col = 1 # A
        end_col = 7   # G

        # Bucle para el borde SUPERIOR
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=start_row, column=col)
            cell.border = Border(top=medium_side, bottom=thin_side, left=cell.border.left, right=cell.border.right)

        # Bucle para el borde INFERIOR
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=end_row, column=col)
            cell.border = Border(top=thin_side, bottom=medium_side, left=cell.border.left, right=cell.border.right)

        # Bucle para el borde IZQUIERDO
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=start_col)
            cell.border = Border(top=cell.border.top, bottom=cell.border.bottom, left=medium_side, right=thin_side)

        # Bucle para el borde DERECHO
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=end_col)
            cell.border = Border(top=cell.border.top, bottom=cell.border.bottom, left=thin_side, right=medium_side)

        # Corregir esquinas que se pisan
        ws.cell(start_row, start_col).border = Border(top=medium_side, left=medium_side, bottom=thin_side, right=thin_side) # A3
        ws.cell(start_row, end_col).border   = Border(top=medium_side, right=medium_side, bottom=thin_side, left=thin_side) # G3
        ws.cell(end_row, start_col).border   = Border(bottom=medium_side, left=medium_side, top=thin_side, right=thin_side) # A(total)
        ws.cell(end_row, end_col).border     = Border(bottom=medium_side, right=medium_side, top=thin_side, left=thin_side) # G(total)

        # 9. Guardar en un archivo temporal
        temp_dir = 'temp_reports'
        os.makedirs(temp_dir, exist_ok=True)
        
        filename = f"Reporte_Sueldos_{month_name}_{year}_{now.strftime('%H%M%S')}.xlsx"
        filepath = os.path.join(temp_dir, filename)
        
        wb.save(filepath)
        
        print(f"Reporte de Excel generado en: {filepath}")
        return filepath, month_name, year
    
    except Exception as e:
        print(f"[Error en excel_generator] No se pudo crear el archivo Excel: {e}")
        return None, None, None